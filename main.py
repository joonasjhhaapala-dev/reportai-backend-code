"""
ReportAI Backend - FastAPI Application
Handles file uploads, data analysis, and report generation
"""

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from typing import Optional
import os
import json
import pandas as pd
import openpyxl
from datetime import datetime
from pathlib import Path

# Import our custom modules
from report_generator import ReportGenerator
from ai_analyzer import AIAnalyzer

# Initialize FastAPI app
app = FastAPI(
    title="ReportAI API",
    description="Automated quality report generation with AI",
    version="1.0.0"
)

# Configure CORS
# Get allowed origins from environment or use defaults
ALLOWED_ORIGINS = os.getenv(
    "ALLOWED_ORIGINS",
    "http://localhost:3000,http://localhost:5173,http://localhost:8000"
).split(",")

# In production, Railway will set this automatically
if os.getenv("RAILWAY_ENVIRONMENT"):
    # Allow Railway frontend URL
    railway_url = os.getenv("RAILWAY_STATIC_URL", "")
    if railway_url:
        ALLOWED_ORIGINS.append(f"https://{railway_url}")
    # Allow any Railway URLs for development
    ALLOWED_ORIGINS.append("https://*.railway.app")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all in production for now (fix later with specific domain)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create directories
UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# Initialize components
report_generator = ReportGenerator()
ai_analyzer = AIAnalyzer()


@app.get("/")
async def root():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "ReportAI Backend",
        "version": "1.0.0"
    }


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    Upload and parse Excel/CSV file
    Returns: File metadata and data preview
    """
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            raise HTTPException(status_code=400, detail="Invalid file type. Only .xlsx, .xls, .csv allowed")
        
        # Save uploaded file
        file_path = UPLOAD_DIR / f"{datetime.now().timestamp()}_{file.filename}"
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Parse file based on type
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file_path)
            sheets = ["Sheet1"]
        else:
            # Read Excel file
            wb = openpyxl.load_workbook(file_path)
            sheets = wb.sheetnames
            # Get first sheet as DataFrame
            df = pd.read_excel(file_path, sheet_name=sheets[0])
        
        # Create data preview
        preview = {
            "rows": len(df),
            "columns": len(df.columns),
            "column_names": df.columns.tolist(),
            "first_rows": df.head(5).to_dict('records'),
            "sheets": sheets
        }
        
        return {
            "success": True,
            "file_id": str(file_path),
            "filename": file.filename,
            "size": len(content),
            "preview": preview
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@app.post("/api/analyze")
async def analyze_data(
    file_id: str = Form(...),
    template_type: str = Form(...),
    language: str = Form(default="en")
):
    """
    Analyze uploaded data with AI
    Returns: AI-generated insights and summary
    """
    try:
        file_path = Path(file_id)
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        
        # Load data
        if file_path.suffix == '.csv':
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        # Analyze with AI
        analysis = await ai_analyzer.analyze(
            data=df,
            template_type=template_type,
            language=language
        )
        
        return {
            "success": True,
            "analysis": analysis
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing data: {str(e)}")


@app.post("/api/generate")
async def generate_report(
    file_id: str = Form(...),
    template_type: str = Form(...),
    report_title: str = Form(...),
    report_date: str = Form(...),
    company_name: str = Form(default=""),
    author_name: str = Form(default=""),
    output_format: str = Form(default="pdf"),
    language: str = Form(default="en")
):
    """
    Generate final report in requested format
    Returns: Download link for generated report
    """
    try:
        file_path = Path(file_id)
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        
        # Load data
        if file_path.suffix == '.csv':
            df = pd.read_csv(file_path)
            wb = None
        else:
            df = pd.read_excel(file_path)
            wb = openpyxl.load_workbook(file_path)
        
        # Get AI analysis
        analysis = await ai_analyzer.analyze(
            data=df,
            template_type=template_type,
            language=language
        )
        
        # Generate report
        report_config = {
            "title": report_title,
            "date": report_date,
            "company": company_name,
            "author": author_name,
            "template_type": template_type,
            "language": language,
            "analysis": analysis
        }
        
        output_file = await report_generator.generate(
            data=df,
            workbook=wb,
            config=report_config,
            output_format=output_format
        )
        
        return {
            "success": True,
            "download_url": f"/api/download/{output_file.name}",
            "filename": output_file.name
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")


@app.get("/api/download/{filename}")
async def download_report(filename: str):
    """
    Download generated report
    """
    file_path = OUTPUT_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type='application/octet-stream'
    )


@app.delete("/api/cleanup/{file_id}")
async def cleanup_files(file_id: str):
    """
    Clean up temporary files
    """
    try:
        file_path = Path(file_id)
        if file_path.exists():
            file_path.unlink()
        return {"success": True, "message": "File cleaned up"}
    except Exception as e:
        return {"success": False, "message": str(e)}


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
